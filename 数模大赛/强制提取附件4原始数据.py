import pandas as pd
from openpyxl import load_workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ======================= 配置区 =======================
INPUT_FILE  = '附件4.xlsx'
OUTPUT_XLSX = '附件4_纯原始数据.xlsx'   # 输出干净的Excel
OUTPUT_CSV  = '附件4_纯原始数据.csv'    # 同时输出一份CSV
# =====================================================

print(f"正在读取 {INPUT_FILE} ...")

# 1. 先用openpyxl打开，看看Excel里所有sheet的真实情况（避免pandas自动读汇总行）
wb = load_workbook(INPUT_FILE, data_only=True)  # data_only=True：如果单元格是公式，只读它显示的数值
print(f"\n📋 工作簿包含以下工作表：{wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- 工作表: {sheet_name} ---")
    print(f"   总行数: {ws.max_row}, 总列数: {ws.max_column}")
    
    # 打印前15行，帮你肉眼判断哪些是原始数据，哪些是汇总行
    print("   前15行内容预览:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15: 
            print(f"   ... (共{ws.max_row}行，后面省略)")
            break
        # 过滤掉全空行
        if any(cell is not None and str(cell).strip() != '' for cell in row):
            print(f"   第{i}行: {row}")

print("\n" + "="*60)
print("开始清理数据：默认把第一个sheet当作原始数据表")
print("如果有多个sheet，请修改脚本里的 sheet_name 参数")
print("="*60)

# 2. 用pandas读取，强制不合并单元格、跳过完全空的行
# 如果你的数据在其他sheet，把下面的 sheet_name=0 改成对应名字，比如 sheet_name='Sheet1'
df = pd.read_excel(
    INPUT_FILE,
    engine='openpyxl',
    sheet_name=0,        # 0=第一个sheet，如果数据在别的sheet就改这里
    header=0,            # 第1行是表头，根据实际调整（比如表头在第2行就写header=1）
    skip_blank_lines=True,
)

# 3. 清理步骤（按需要注释/取消注释）
# --- 3.1 删除整行都是空的
df = df.dropna(how='all')

# --- 3.2 删除整列都是空的
df = df.dropna(axis=1, how='all')

# --- 3.3 关键：去掉"汇总行/总计行/平均值行"
# 分类汇总的汇总行通常在第一列是空白，或者含"汇总/总计/平均"字样
# 如果第一列有分类编码，汇总行的编码列通常是空的 → 删掉
if df.shape[1] >= 1:
    first_col = df.columns[0]
    before = len(df)
    # 删掉第一列为空的行（汇总行通常在第一列没有编码）
    df = df[df[first_col].notna()]
    # 再删掉第一列包含"汇总/总计/平均/合计"的行
    mask_bad = df[first_col].astype(str).str.contains('汇总|总计|平均|合计|Grand Total', na=False)
    df = df[~mask_bad]
    after = len(df)
    if before != after:
        print(f"✅ 已删除 {before - after} 行可能的汇总行/平均值行")

# --- 3.4 如果某列是数字，但被读成文本了，强制转一下（按需开启）
# for col in df.columns:
#     df[col] = pd.to_numeric(df[col], errors='ignore')

# 4. 打印最终结果
print(f"\n✅ 清理完成：最终保留 {len(df)} 行, {len(df.columns)} 列")
print("列名：", df.columns.tolist())
print("\n📊 最终数据预览（前20行）：")
print(df.to_string())

# 5. 保存干净的文件
df.to_excel(OUTPUT_XLSX, index=False)
df.to_csv(OUTPUT_CSV, index=False, encoding='utf-8-sig')
print(f"\n💾 已保存到:")
print(f"   Excel → {OUTPUT_XLSX}")
print(f"   CSV   → {OUTPUT_CSV}")
print("\n🎉 搞定！现在这两个文件就是纯原始数据，没有任何平均值汇总了。")